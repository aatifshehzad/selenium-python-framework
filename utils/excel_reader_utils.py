import openpyxl

from utils.logger import Logger


class ExcelReader:
    obj_logger = Logger().get_logger()

    def __get_excel_sheet_by_name(self, str_excel_path, str_sheet_name):
        try:
            excel_book = openpyxl.load_workbook(str_excel_path)
            if str_sheet_name in excel_book.sheetnames:
                active_sheet = excel_book[str_sheet_name]
                return active_sheet
            else:
                raise Exception("Provide Valid Sheet Name.")
        except Exception as e:
            self.obj_logger.error("An Exception Occurred: " + str(e))
            raise Exception(e)

    def get_value_by_index(self, str_excel_path, str_sheet_name, int_row, int_column):
        try:
            active_sheet = self.__get_excel_sheet_by_name(str_excel_path, str_sheet_name)
            return active_sheet.cell(row=int_row, column=int_column).value
        except Exception as e:
            self.obj_logger.error("An Exception Occurred: " + str(e))
            raise Exception(e)

    def get_data_in_list_of_dict(self, str_excel_path, str_sheet_name):
        try:
            data_list = []
            active_sheet = self.__get_excel_sheet_by_name(str_excel_path, str_sheet_name)
            int_max_rows = active_sheet.max_row
            int_max_columns = active_sheet.max_column

            for int_row in range(2, int_max_rows + 1):
                data_dict = {}
                for int_column in range(1, int_max_columns + 1):
                    data_dict[active_sheet.cell(row=1, column=int_column).value] = active_sheet.cell(row=int_row,
                                                                                                     column=int_column).value
                data_list.append(data_dict)
            return data_list
        except Exception as e:
            self.obj_logger.error("An Exception Occurred: " + str(e))
            raise Exception(e)
